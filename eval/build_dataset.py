"""Builds the 300-entry intent classification dataset and exports to XLSX with dropdowns."""
import json
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

NEW_ENTRIES = [
  # ── content_generation ──────────────────────────────────────────────────────
  {"id":151,"query":"I need a lesson on the human digestive system for Class 8 biology","translated_query":"I need a lesson on the human digestive system for Class 8 biology","intent":"content_generation","language":"en"},
  {"id":152,"query":"Design an activity to teach place value in large numbers for Class 5","translated_query":"Design an activity to teach place value in large numbers for Class 5","intent":"content_generation","language":"en"},
  {"id":153,"query":"Give me content on types of soil and their properties for Class 7 geography","translated_query":"Give me content on types of soil and their properties for Class 7 geography","intent":"content_generation","language":"en"},
  {"id":154,"query":"Create a lesson on light and reflection for Class 10 physics","translated_query":"Create a lesson on light and reflection for Class 10 physics","intent":"content_generation","language":"en"},
  {"id":155,"query":"Prepare a Class 7 history lesson on the Mughal Empire with a fun activity","translated_query":"Prepare a Class 7 history lesson on the Mughal Empire with a fun activity","intent":"content_generation","language":"en"},
  {"id":156,"query":"Class 9 civics — I need an engaging lesson on fundamental rights","translated_query":"Class 9 civics — I need an engaging lesson on fundamental rights","intent":"content_generation","language":"en"},

  {"id":157,"query":"कक्षा 8 के लिए पाचन तंत्र का पाठ बनाओ","translated_query":"Make a lesson on the digestive system for Class 8","intent":"content_generation","language":"hi"},
  {"id":158,"query":"कक्षा 5 में बड़ी संख्याओं का स्थानीय मान सिखाने के लिए गतिविधि दो","translated_query":"Give an activity to teach place value of large numbers in Class 5","intent":"content_generation","language":"hi"},
  {"id":159,"query":"कक्षा 7 भूगोल के लिए मिट्टी के प्रकार और उनके गुणों का कंटेंट दो","translated_query":"Give content on types of soil and their properties for Class 7 geography","intent":"content_generation","language":"hi"},
  {"id":160,"query":"कक्षा 10 भौतिक विज्ञान में प्रकाश और परावर्तन का पाठ बनाइए","translated_query":"Create a lesson on light and reflection for Class 10 physics","intent":"content_generation","language":"hi"},
  {"id":161,"query":"कक्षा 7 इतिहास में मुगल साम्राज्य पर पाठ बनाओ, कोई मजेदार activity भी हो","translated_query":"Make a lesson on the Mughal Empire for Class 7 history, include a fun activity","intent":"content_generation","language":"hi"},
  {"id":162,"query":"कक्षा 9 नागरिक शास्त्र के लिए मौलिक अधिकारों का पाठ चाहिए","translated_query":"Need a lesson on fundamental rights for Class 9 civics","intent":"content_generation","language":"hi"},

  {"id":163,"query":"Class 8 biology mein digestive system ka lesson do","translated_query":"Give a digestive system lesson for Class 8 biology","intent":"content_generation","language":"hinglish"},
  {"id":164,"query":"Class 5 ke liye badi sankhyaon ka sthaniy man sikhane ki activity chahiye","translated_query":"Need an activity to teach place value of large numbers for Class 5","intent":"content_generation","language":"hinglish"},
  {"id":165,"query":"Soil types ke baare mein Class 7 geography ka content do, properties bhi batao","translated_query":"Give Class 7 geography content on soil types, mention properties too","intent":"content_generation","language":"hinglish"},
  {"id":166,"query":"Class 10 physics mein light aur reflection ka lesson banao","translated_query":"Make a light and reflection lesson for Class 10 physics","intent":"content_generation","language":"hinglish"},
  {"id":167,"query":"Mughal Empire pe Class 7 history ka lesson chahiye, kuch interesting activity ke saath","translated_query":"Need a Class 7 history lesson on the Mughal Empire with an interesting activity","intent":"content_generation","language":"hinglish"},
  {"id":168,"query":"Class 9 civics ke liye fundamental rights pe engaging lesson do","translated_query":"Give an engaging lesson on fundamental rights for Class 9 civics","intent":"content_generation","language":"hinglish"},

  {"id":169,"query":"வகுப்பு 8 உயிரியல் செரிமான மண்டலம் பாடம் தயாரி","translated_query":"Prepare a digestive system lesson for Class 8 biology","intent":"content_generation","language":"ta"},
  {"id":170,"query":"வகுப்பு 5 க்கு பெரிய எண்களின் இட மதிப்பு செயல்பாடு கொடு","translated_query":"Give a place value activity for large numbers for Class 5","intent":"content_generation","language":"ta"},
  {"id":171,"query":"வகுப்பு 7 புவியியல் மண் வகைகள் மற்றும் பண்புகள் பாடம்","translated_query":"Lesson on soil types and properties for Class 7 geography","intent":"content_generation","language":"ta"},
  {"id":172,"query":"வகுப்பு 10 இயற்பியல் ஒளி மற்றும் பிரதிபலிப்பு பாடம் வேண்டும்","translated_query":"Need a lesson on light and reflection for Class 10 physics","intent":"content_generation","language":"ta"},
  {"id":173,"query":"வகுப்பு 7 வரலாறு முகலாயப் பேரரசு பாடம் சுவாரஸ்யமாக","translated_query":"Make the Class 7 history Mughal Empire lesson interesting","intent":"content_generation","language":"ta"},
  {"id":174,"query":"வகுப்பு 9 குடிமையியல் அடிப்படை உரிமைகள் பாடம் வேண்டும்","translated_query":"Need a lesson on fundamental rights for Class 9 civics","intent":"content_generation","language":"ta"},

  {"id":175,"query":"తరగతి 8 జీవశాస్త్రానికి జీర్ణ వ్యవస్థ పాఠం తయారుచేయి","translated_query":"Prepare a digestive system lesson for Class 8 biology","intent":"content_generation","language":"te"},
  {"id":176,"query":"తరగతి 5 కోసం పెద్ద సంఖ్యల స్థాన విలువ కార్యకలాపం ఇవ్వు","translated_query":"Give a place value activity for large numbers for Class 5","intent":"content_generation","language":"te"},
  {"id":177,"query":"తరగతి 7 భూగోళ శాస్త్రానికి నేల రకాలు మరియు వాటి లక్షణాల పాఠం","translated_query":"Lesson on soil types and their properties for Class 7 geography","intent":"content_generation","language":"te"},
  {"id":178,"query":"తరగతి 10 భౌతిక శాస్త్రానికి కాంతి మరియు పరావర్తనంపై పాఠం","translated_query":"Lesson on light and reflection for Class 10 physics","intent":"content_generation","language":"te"},
  {"id":179,"query":"తరగతి 7 చరిత్రకు మొఘల్ సామ్రాజ్యంపై ఆసక్తికరమైన పాఠం","translated_query":"Interesting lesson on the Mughal Empire for Class 7 history","intent":"content_generation","language":"te"},
  {"id":180,"query":"తరగతి 9 పౌరశాస్త్రానికి ప్రాథమిక హక్కులపై పాఠం కావాలి","translated_query":"Need a lesson on fundamental rights for Class 9 civics","intent":"content_generation","language":"te"},

  # ── feedback ────────────────────────────────────────────────────────────────
  {"id":181,"query":"The hook was great but the example didn't land with my students at all","translated_query":"The hook was great but the example didn't land with my students at all","intent":"feedback","language":"en"},
  {"id":182,"query":"Response is useful but please remove the emoji formatting, it looks unprofessional","translated_query":"Response is useful but please remove the emoji formatting, it looks unprofessional","intent":"feedback","language":"en"},
  {"id":183,"query":"Can you generate content suited for a mixed-ability classroom?","translated_query":"Can you generate content suited for a mixed-ability classroom?","intent":"feedback","language":"en"},
  {"id":184,"query":"First time using ShikshaBot and I'm genuinely impressed. Very useful tool.","translated_query":"First time using ShikshaBot and I'm genuinely impressed. Very useful tool.","intent":"feedback","language":"en"},
  {"id":185,"query":"The SEL section felt forced and unnatural, it didn't connect to the topic","translated_query":"The SEL section felt forced and unnatural, it didn't connect to the topic","intent":"feedback","language":"en"},
  {"id":186,"query":"I asked for Class 6 content but the difficulty level was more like Class 8","translated_query":"I asked for Class 6 content but the difficulty level was more like Class 8","intent":"feedback","language":"en"},

  {"id":187,"query":"Hook तो अच्छा था लेकिन उदाहरण बच्चों को समझ नहीं आया","translated_query":"The hook was good but the example didn't work for the children","intent":"feedback","language":"hi"},
  {"id":188,"query":"Response उपयोगी है लेकिन emoji हटाओ, professional नहीं लगता","translated_query":"The response is useful but remove the emojis, it doesn't look professional","intent":"feedback","language":"hi"},
  {"id":189,"query":"Mixed ability class के लिए content बना सकते हो क्या?","translated_query":"Can you make content for a mixed ability class?","intent":"feedback","language":"hi"},
  {"id":190,"query":"पहली बार ShikshaBot use किया, बहुत अच्छा लगा, काम का है","translated_query":"Used ShikshaBot for the first time, liked it a lot, it's useful","intent":"feedback","language":"hi"},
  {"id":191,"query":"SEL वाला हिस्सा artificial लगा, topic से naturally नहीं जुड़ा","translated_query":"The SEL part felt artificial, didn't connect naturally to the topic","intent":"feedback","language":"hi"},
  {"id":192,"query":"Class 6 माँगा था लेकिन content Class 8 level का था","translated_query":"I asked for Class 6 but the content was at Class 8 level","intent":"feedback","language":"hi"},

  {"id":193,"query":"Hook achha tha par example students ko bilkul samajh nahi aaya","translated_query":"The hook was good but the students didn't understand the example at all","intent":"feedback","language":"hinglish"},
  {"id":194,"query":"Content theek hai lekin emoji mat dalo, professional nahi lagta","translated_query":"Content is fine but don't add emojis, it doesn't look professional","intent":"feedback","language":"hinglish"},
  {"id":195,"query":"Kya mixed-ability class ke liye bhi content de sakte ho?","translated_query":"Can you also give content for a mixed-ability class?","intent":"feedback","language":"hinglish"},
  {"id":196,"query":"Pehli baar ShikshaBot use kiya, bahut impressed hoon, useful hai","translated_query":"Used ShikshaBot for the first time, very impressed, it's useful","intent":"feedback","language":"hinglish"},
  {"id":197,"query":"SEL wala part forced lag raha tha, topic se naturally connect nahi hua","translated_query":"The SEL part felt forced, didn't connect naturally to the topic","intent":"feedback","language":"hinglish"},
  {"id":198,"query":"Maine Class 6 manga tha lekin content Class 8 level ka tha","translated_query":"I asked for Class 6 but the content was at Class 8 level","intent":"feedback","language":"hinglish"},

  {"id":199,"query":"Hook நன்றாக இருந்தது ஆனால் உதாரணம் மாணவர்களுக்கு புரியவில்லை","translated_query":"The hook was good but the example was not understood by students","intent":"feedback","language":"ta"},
  {"id":200,"query":"பதில் பயனுள்ளது ஆனால் emoji வேண்டாம், தொழில்முறையாக தெரியவில்லை","translated_query":"The response is useful but no emojis, it doesn't look professional","intent":"feedback","language":"ta"},
  {"id":201,"query":"Mixed ability வகுப்பிற்கு content தர முடியுமா?","translated_query":"Can you give content for a mixed ability class?","intent":"feedback","language":"ta"},
  {"id":202,"query":"முதல் முறை ShikshaBot பயன்படுத்தினேன், மிகவும் பிடித்தது","translated_query":"Used ShikshaBot for the first time, liked it very much","intent":"feedback","language":"ta"},
  {"id":203,"query":"SEL பகுதி இயற்கையாக இல்லை, தலைப்புடன் இணைக்கவில்லை","translated_query":"The SEL part was not natural, didn't connect with the topic","intent":"feedback","language":"ta"},
  {"id":204,"query":"வகுப்பு 6 கேட்டேன் ஆனால் வகுப்பு 8 அளவிலான content கொடுத்தீர்கள்","translated_query":"I asked for Class 6 but you gave Class 8 level content","intent":"feedback","language":"ta"},

  {"id":205,"query":"హుక్ బాగుంది కానీ ఉదాహరణ విద్యార్థులకు అర్థం కాలేదు","translated_query":"The hook was good but the example was not understood by students","intent":"feedback","language":"te"},
  {"id":206,"query":"సమాధానం ఉపయోగకరంగా ఉంది కానీ ఎమోజీలు వద్దు, వృత్తిపరంగా లేదు","translated_query":"The response is useful but no emojis, it doesn't look professional","intent":"feedback","language":"te"},
  {"id":207,"query":"మిక్స్డ్ అబిలిటీ తరగతికి కంటెంట్ ఇవ్వగలవా?","translated_query":"Can you give content for a mixed ability class?","intent":"feedback","language":"te"},
  {"id":208,"query":"మొదటిసారి ShikshaBot వాడాను, చాలా నచ్చింది, ఉపయోగకరంగా ఉంది","translated_query":"Used ShikshaBot for the first time, liked it a lot, it's useful","intent":"feedback","language":"te"},
  {"id":209,"query":"SEL భాగం సహజంగా అనిపించలేదు, అంశంతో అనుసంధానం కాలేదు","translated_query":"The SEL part didn't feel natural, didn't connect with the topic","intent":"feedback","language":"te"},
  {"id":210,"query":"తరగతి 6 అడిగాను కానీ కంటెంట్ తరగతి 8 స్థాయిలో ఉంది","translated_query":"I asked for Class 6 but the content was at Class 8 level","intent":"feedback","language":"te"},

  # ── query_resolution_academic ────────────────────────────────────────────────
  {"id":211,"query":"What is the difference between speed and velocity?","translated_query":"What is the difference between speed and velocity?","intent":"query_resolution_academic","language":"en"},
  {"id":212,"query":"How do vaccines work in the human body?","translated_query":"How do vaccines work in the human body?","intent":"query_resolution_academic","language":"en"},
  {"id":213,"query":"What is the Preamble of the Indian Constitution?","translated_query":"What is the Preamble of the Indian Constitution?","intent":"query_resolution_academic","language":"en"},
  {"id":214,"query":"What is a food web and how is it different from a food chain?","translated_query":"What is a food web and how is it different from a food chain?","intent":"query_resolution_academic","language":"en"},
  {"id":215,"query":"What were the effects of the Industrial Revolution on society?","translated_query":"What were the effects of the Industrial Revolution on society?","intent":"query_resolution_academic","language":"en"},
  {"id":216,"query":"Explain the process of rusting with the chemical equation","translated_query":"Explain the process of rusting with the chemical equation","intent":"query_resolution_academic","language":"en"},

  {"id":217,"query":"चाल और वेग में क्या अंतर है?","translated_query":"What is the difference between speed and velocity?","intent":"query_resolution_academic","language":"hi"},
  {"id":218,"query":"टीका मानव शरीर में कैसे काम करता है?","translated_query":"How does a vaccine work in the human body?","intent":"query_resolution_academic","language":"hi"},
  {"id":219,"query":"भारतीय संविधान की प्रस्तावना क्या है?","translated_query":"What is the Preamble of the Indian Constitution?","intent":"query_resolution_academic","language":"hi"},
  {"id":220,"query":"खाद्य जाल क्या है और यह खाद्य श्रृंखला से कैसे अलग है?","translated_query":"What is a food web and how is it different from a food chain?","intent":"query_resolution_academic","language":"hi"},
  {"id":221,"query":"औद्योगिक क्रांति के समाज पर क्या प्रभाव पड़े?","translated_query":"What were the effects of the Industrial Revolution on society?","intent":"query_resolution_academic","language":"hi"},
  {"id":222,"query":"जंग लगने की प्रक्रिया रासायनिक समीकरण के साथ समझाइए","translated_query":"Explain the process of rusting with the chemical equation","intent":"query_resolution_academic","language":"hi"},

  {"id":223,"query":"Speed aur velocity mein kya fark hai, simply batao","translated_query":"What is the difference between speed and velocity, explain simply","intent":"query_resolution_academic","language":"hinglish"},
  {"id":224,"query":"Vaccine human body mein kaise kaam karta hai?","translated_query":"How does a vaccine work in the human body?","intent":"query_resolution_academic","language":"hinglish"},
  {"id":225,"query":"Indian Constitution ki Preamble kya kehti hai?","translated_query":"What does the Preamble of the Indian Constitution say?","intent":"query_resolution_academic","language":"hinglish"},
  {"id":226,"query":"Food web kya hota hai? Food chain se kaise different hai?","translated_query":"What is a food web? How is it different from a food chain?","intent":"query_resolution_academic","language":"hinglish"},
  {"id":227,"query":"Industrial Revolution ke kya effects the society pe?","translated_query":"What were the effects of the Industrial Revolution on society?","intent":"query_resolution_academic","language":"hinglish"},
  {"id":228,"query":"Rusting ki chemical process kya hai, equation bhi batao","translated_query":"What is the chemical process of rusting, also give the equation","intent":"query_resolution_academic","language":"hinglish"},

  {"id":229,"query":"வேகம் மற்றும் திசைவேகம் இடையே என்ன வித்தியாசம்?","translated_query":"What is the difference between speed and velocity?","intent":"query_resolution_academic","language":"ta"},
  {"id":230,"query":"தடுப்பூசி மனித உடலில் எவ்வாறு செயல்படுகிறது?","translated_query":"How does a vaccine work in the human body?","intent":"query_resolution_academic","language":"ta"},
  {"id":231,"query":"இந்திய அரசியலமைப்பின் முன்னுரை என்ன சொல்கிறது?","translated_query":"What does the Preamble of the Indian Constitution say?","intent":"query_resolution_academic","language":"ta"},
  {"id":232,"query":"உணவு வலை என்றால் என்ன? உணவு சங்கிலியிலிருந்து எவ்வாறு வேறுபடுகிறது?","translated_query":"What is a food web? How does it differ from a food chain?","intent":"query_resolution_academic","language":"ta"},
  {"id":233,"query":"தொழில்துறை புரட்சியின் சமூகத்தில் விளைவுகள் என்ன?","translated_query":"What were the effects of the Industrial Revolution on society?","intent":"query_resolution_academic","language":"ta"},
  {"id":234,"query":"துருப்பிடிக்கும் வேதியியல் செயல்முறையை சமன்பாட்டுடன் விளக்குங்கள்","translated_query":"Explain the chemical process of rusting with the equation","intent":"query_resolution_academic","language":"ta"},

  {"id":235,"query":"వేగం మరియు వేగం (velocity) మధ్య తేడా ఏమిటి?","translated_query":"What is the difference between speed and velocity?","intent":"query_resolution_academic","language":"te"},
  {"id":236,"query":"వ్యాక్సిన్ మానవ శరీరంలో ఎలా పని చేస్తుంది?","translated_query":"How does a vaccine work in the human body?","intent":"query_resolution_academic","language":"te"},
  {"id":237,"query":"భారత రాజ్యాంగ ప్రవేశిక ఏమి చెప్తుంది?","translated_query":"What does the Preamble of the Indian Constitution say?","intent":"query_resolution_academic","language":"te"},
  {"id":238,"query":"ఆహార వల అంటే ఏమిటి? ఆహార గొలుసు నుండి ఎలా వేరుగా ఉంటుంది?","translated_query":"What is a food web? How is it different from a food chain?","intent":"query_resolution_academic","language":"te"},
  {"id":239,"query":"పారిశ్రామిక విప్లవం సమాజంపై ఏ ప్రభావాలు చూపింది?","translated_query":"What effects did the Industrial Revolution have on society?","intent":"query_resolution_academic","language":"te"},
  {"id":240,"query":"తుప్పు పట్టే రసాయన ప్రక్రియను సమీకరణంతో వివరించండి","translated_query":"Explain the chemical process of rusting with the equation","intent":"query_resolution_academic","language":"te"},

  # ── query_resolution_sel ─────────────────────────────────────────────────────
  {"id":241,"query":"A student who was always a topper has started failing and seems very withdrawn lately","translated_query":"A student who was always a topper has started failing and seems very withdrawn lately","intent":"query_resolution_sel","language":"en"},
  {"id":242,"query":"After being absent for two weeks, my student came back very sad and doesn't talk to anyone","translated_query":"After being absent for two weeks, my student came back very sad and doesn't talk to anyone","intent":"query_resolution_sel","language":"en"},
  {"id":243,"query":"Some students are forming cliques and excluding one child from all group activities","translated_query":"Some students are forming cliques and excluding one child from all group activities","intent":"query_resolution_sel","language":"en"},
  {"id":244,"query":"One student always refuses to work in groups and insists on doing everything alone","translated_query":"One student always refuses to work in groups and insists on doing everything alone","intent":"query_resolution_sel","language":"en"},
  {"id":245,"query":"I've noticed a student coming to school hungry every day, they seem distracted all the time","translated_query":"I've noticed a student coming to school hungry every day, they seem distracted all the time","intent":"query_resolution_sel","language":"en"},
  {"id":246,"query":"A child keeps saying things like 'I'm too stupid for this' and 'I can never do anything right'","translated_query":"A child keeps saying things like 'I'm too stupid for this' and 'I can never do anything right'","intent":"query_resolution_sel","language":"en"},

  {"id":247,"query":"जो बच्चा हमेशा first आता था वो अब fail हो रहा है और बहुत withdrawn हो गया है","translated_query":"A child who always came first is now failing and has become very withdrawn","intent":"query_resolution_sel","language":"hi"},
  {"id":248,"query":"दो हफ्ते absent रहने के बाद बच्चा बहुत उदास लौटा है और किसी से बात नहीं करता","translated_query":"After two weeks of absence the child came back very sad and doesn't talk to anyone","intent":"query_resolution_sel","language":"hi"},
  {"id":249,"query":"कुछ बच्चे group बना कर एक बच्चे को सभी activities से बाहर रख रहे हैं","translated_query":"Some children are forming groups and excluding one child from all activities","intent":"query_resolution_sel","language":"hi"},
  {"id":250,"query":"एक बच्चा कभी भी group में काम नहीं करता, हमेशा अकेले करना चाहता है","translated_query":"One child never works in a group, always wants to do things alone","intent":"query_resolution_sel","language":"hi"},
  {"id":251,"query":"एक बच्चा रोज भूखा स्कूल आता है, पूरे दिन उसका ध्यान नहीं लगता","translated_query":"A child comes to school hungry every day, can't concentrate all day","intent":"query_resolution_sel","language":"hi"},
  {"id":252,"query":"एक बच्चा बार बार कहता है 'मैं बहुत stupid हूँ', खुद को बेकार समझता है","translated_query":"A child keeps saying 'I am very stupid', considers themselves useless","intent":"query_resolution_sel","language":"hi"},

  {"id":253,"query":"Jo baccha pehle topper tha ab fail ho raha hai aur bahut withdrawn ho gaya hai","translated_query":"A child who was previously a topper is now failing and has become very withdrawn","intent":"query_resolution_sel","language":"hinglish"},
  {"id":254,"query":"Do hafte baad aaya to bahut udaas tha, kisi se baat nahi karta ab","translated_query":"After two weeks came back very sad, doesn't talk to anyone now","intent":"query_resolution_sel","language":"hinglish"},
  {"id":255,"query":"Kuch bachche group bana kar ek bacche ko activities se bahar rakh rahe hain","translated_query":"Some children are forming groups and excluding one child from activities","intent":"query_resolution_sel","language":"hinglish"},
  {"id":256,"query":"Ek baccha kabhi group mein kaam nahi karta, hamesha akele karna chahta hai","translated_query":"One child never works in a group, always wants to do things alone","intent":"query_resolution_sel","language":"hinglish"},
  {"id":257,"query":"Ek baccha roz bhooka school aata hai, din bhar concentrate nahi kar pata","translated_query":"A child comes to school hungry every day, can't concentrate throughout the day","intent":"query_resolution_sel","language":"hinglish"},
  {"id":258,"query":"Ek baccha baar baar kehta hai 'main bahut stupid hoon', khud ko bekar samajhta hai","translated_query":"A child keeps saying 'I am very stupid', considers themselves worthless","intent":"query_resolution_sel","language":"hinglish"},

  {"id":259,"query":"எப்போதும் முதலிடத்தில் இருந்த மாணவன் தோல்வி அடைகிறான், மிகவும் தனிமையாகி விட்டான்","translated_query":"A student who always topped is now failing and has become very withdrawn","intent":"query_resolution_sel","language":"ta"},
  {"id":260,"query":"இரண்டு வாரங்கள் இல்லாமல் வந்த மாணவன் மிகவும் சோகமாக வந்தான், யாரிடமும் பேசவில்லை","translated_query":"After two weeks of absence the student came back very sad and doesn't talk to anyone","intent":"query_resolution_sel","language":"ta"},
  {"id":261,"query":"சில மாணவர்கள் குழு உருவாக்கி ஒரு குழந்தையை எல்லா செயல்பாடுகளிலிருந்தும் ஒதுக்குகிறார்கள்","translated_query":"Some students are forming groups and excluding one child from all activities","intent":"query_resolution_sel","language":"ta"},
  {"id":262,"query":"ஒரு மாணவன் குழு வேலையை மறுக்கிறான், எப்போதும் தனியாக செய்ய விரும்புகிறான்","translated_query":"A student refuses group work, always wants to do things alone","intent":"query_resolution_sel","language":"ta"},
  {"id":263,"query":"ஒரு குழந்தை தினமும் பசியுடன் பள்ளிக்கு வருகிறது, நாள் முழுவதும் கவனிக்கவில்லை","translated_query":"A child comes to school hungry every day, doesn't concentrate all day","intent":"query_resolution_sel","language":"ta"},
  {"id":264,"query":"ஒரு குழந்தை தொடர்ந்து 'நான் மிகவும் முட்டாள்' என்று சொல்கிறது","translated_query":"A child keeps saying 'I am very stupid'","intent":"query_resolution_sel","language":"ta"},

  {"id":265,"query":"ఎప్పుడూ టాపర్‌గా ఉన్న విద్యార్థి ఇప్పుడు ఫెయిల్ అవుతున్నాడు మరియు చాలా దూరంగా ఉన్నాడు","translated_query":"A student who was always a topper is now failing and has become very withdrawn","intent":"query_resolution_sel","language":"te"},
  {"id":266,"query":"రెండు వారాల తర్వాత వచ్చిన విద్యార్థి చాలా దుఃఖంగా ఉన్నాడు, ఎవరితోనూ మాట్లాడడు","translated_query":"After two weeks of absence the student came back very sad and doesn't talk to anyone","intent":"query_resolution_sel","language":"te"},
  {"id":267,"query":"కొంతమంది విద్యార్థులు గుంపులు చేసుకుని ఒక పిల్లవాడిని అన్ని కార్యకలాపాల నుండి మినహాయిస్తున్నారు","translated_query":"Some students are forming groups and excluding one child from all activities","intent":"query_resolution_sel","language":"te"},
  {"id":268,"query":"ఒక విద్యార్థి గ్రూప్ వర్క్ చేయడానికి ఎప్పుడూ నిరాకరిస్తాడు, ఒంటరిగా చేయాలని ఉంటాడు","translated_query":"A student always refuses group work, wants to do things alone","intent":"query_resolution_sel","language":"te"},
  {"id":269,"query":"ఒక పిల్లవాడు రోజూ పాఠశాలకు ఆకలిగా వస్తున్నాడు, రోజంతా ఏకాగ్రత లేదు","translated_query":"A child comes to school hungry every day, no concentration throughout the day","intent":"query_resolution_sel","language":"te"},
  {"id":270,"query":"ఒక పిల్లవాడు పదే పదే 'నేను చాలా stupid' అని చెప్తాడు, తనను తాను నిందించుకుంటాడు","translated_query":"A child keeps saying 'I am very stupid' and blames themselves","intent":"query_resolution_sel","language":"te"},

  # ── out_of_service ───────────────────────────────────────────────────────────
  {"id":271,"query":"What diet should I follow to lose belly fat quickly?","translated_query":"What diet should I follow to lose belly fat quickly?","intent":"out_of_service","language":"en"},
  {"id":272,"query":"Can you write a wedding speech for my brother's marriage next week?","translated_query":"Can you write a wedding speech for my brother's marriage next week?","intent":"out_of_service","language":"en"},
  {"id":273,"query":"Help me draft a complaint letter to the housing society about water supply","translated_query":"Help me draft a complaint letter to the housing society about water supply","intent":"out_of_service","language":"en"},
  {"id":274,"query":"What is the current gold price in India today?","translated_query":"What is the current gold price in India today?","intent":"out_of_service","language":"en"},
  {"id":275,"query":"Suggest a good Netflix series to watch this weekend with family","translated_query":"Suggest a good Netflix series to watch this weekend with family","intent":"out_of_service","language":"en"},
  {"id":276,"query":"My phone screen cracked. Should I repair it or buy a new phone?","translated_query":"My phone screen cracked. Should I repair it or buy a new phone?","intent":"out_of_service","language":"en"},

  {"id":277,"query":"पेट की चर्बी कम करने के लिए कौन सा diet follow करूँ?","translated_query":"Which diet should I follow to reduce belly fat?","intent":"out_of_service","language":"hi"},
  {"id":278,"query":"अगले हफ्ते भाई की शादी है, speech लिख दो","translated_query":"My brother's wedding is next week, write a speech","intent":"out_of_service","language":"hi"},
  {"id":279,"query":"पानी की supply के बारे में housing society को complaint letter draft करो","translated_query":"Draft a complaint letter to the housing society about water supply","intent":"out_of_service","language":"hi"},
  {"id":280,"query":"आज भारत में सोने का भाव क्या है?","translated_query":"What is the gold price in India today?","intent":"out_of_service","language":"hi"},
  {"id":281,"query":"इस weekend family के साथ देखने के लिए कोई Netflix series बताओ","translated_query":"Suggest a Netflix series to watch with family this weekend","intent":"out_of_service","language":"hi"},
  {"id":282,"query":"मेरे फोन की screen टूट गई, repair करवाऊँ या नया लूँ?","translated_query":"My phone screen broke, should I repair it or buy a new one?","intent":"out_of_service","language":"hi"},

  {"id":283,"query":"Belly fat khatam karne ke liye kaunsa diet follow karoon?","translated_query":"Which diet should I follow to lose belly fat?","intent":"out_of_service","language":"hinglish"},
  {"id":284,"query":"Bhai ki shaadi hai agle hafte, wedding speech likh do","translated_query":"My brother's wedding is next week, write a wedding speech","intent":"out_of_service","language":"hinglish"},
  {"id":285,"query":"Paani ki problem ke baare mein housing society ko complaint letter chahiye","translated_query":"Need a complaint letter to the housing society about the water problem","intent":"out_of_service","language":"hinglish"},
  {"id":286,"query":"Aaj sone ka bhav kya hai India mein?","translated_query":"What is the gold price in India today?","intent":"out_of_service","language":"hinglish"},
  {"id":287,"query":"Is weekend family ke saath dekhne ke liye Netflix pe kya hai achha?","translated_query":"What's good on Netflix to watch with family this weekend?","intent":"out_of_service","language":"hinglish"},
  {"id":288,"query":"Meri phone ki screen toot gayi, repair karwaaun ya naya loon?","translated_query":"My phone screen broke, should I get it repaired or buy a new one?","intent":"out_of_service","language":"hinglish"},

  {"id":289,"query":"வயிற்று கொழுப்பை குறைக்க என்ன diet பின்பற்றவேண்டும்?","translated_query":"What diet should I follow to reduce belly fat?","intent":"out_of_service","language":"ta"},
  {"id":290,"query":"அடுத்த வாரம் என் அண்ணன் திருமணம், பேச்சு எழுதி கொடுங்கள்","translated_query":"My brother's wedding is next week, please write a speech","intent":"out_of_service","language":"ta"},
  {"id":291,"query":"தண்ணீர் விநியோக பிரச்சனையில் வீட்டு சங்கத்திற்கு புகார் கடிதம் தேவை","translated_query":"Need a complaint letter to the housing society about the water supply problem","intent":"out_of_service","language":"ta"},
  {"id":292,"query":"இன்று இந்தியாவில் தங்கத்தின் விலை என்ன?","translated_query":"What is the gold price in India today?","intent":"out_of_service","language":"ta"},
  {"id":293,"query":"இந்த வார இறுதியில் குடும்பத்துடன் பார்க்க Netflix-ல் என்ன நல்லது?","translated_query":"What's good on Netflix to watch with family this weekend?","intent":"out_of_service","language":"ta"},
  {"id":294,"query":"என் ஃபோன் திரை உடைந்துவிட்டது, பழுது பார்க்கலாமா அல்லது புதியது வாங்கலாமா?","translated_query":"My phone screen broke, should I repair it or buy a new one?","intent":"out_of_service","language":"ta"},

  {"id":295,"query":"పొట్ట కొవ్వు తగ్గించుకోవడానికి ఏ డైట్ అనుసరించాలి?","translated_query":"Which diet should I follow to reduce belly fat?","intent":"out_of_service","language":"te"},
  {"id":296,"query":"వచ్చే వారం నా అన్న పెళ్లి, వివాహ ప్రసంగం రాయండి","translated_query":"My brother's wedding is next week, write a wedding speech","intent":"out_of_service","language":"te"},
  {"id":297,"query":"నీటి సరఫరా సమస్యపై హౌసింగ్ సొసైటీకి ఫిర్యాదు లేఖ కావాలి","translated_query":"Need a complaint letter to the housing society about the water supply issue","intent":"out_of_service","language":"te"},
  {"id":298,"query":"ఈరోజు భారతదేశంలో బంగారం ధర ఎంత?","translated_query":"What is the gold price in India today?","intent":"out_of_service","language":"te"},
  {"id":299,"query":"ఈ వారాంతంలో కుటుంబంతో చూడటానికి Netflix లో ఏమి బాగుంటుంది?","translated_query":"What's good on Netflix to watch with family this weekend?","intent":"out_of_service","language":"te"},
  {"id":300,"query":"నా ఫోన్ స్క్రీన్ పగిలిపోయింది, రిపేర్ చేయించాలా లేదా కొత్తది కొనాలా?","translated_query":"My phone screen broke, should I repair it or buy a new one?","intent":"out_of_service","language":"te"},
]

INTENTS = [
    "content_generation",
    "feedback",
    "query_resolution_academic",
    "query_resolution_sel",
    "out_of_service",
]

LANG_COLORS = {
    "en":       "DDEEFF",
    "hi":       "FFE5CC",
    "hinglish": "E8F5E9",
    "ta":       "FFF9C4",
    "te":       "F3E5F5",
}

def load_existing():
    with open("intent_classification_dataset.json") as f:
        return json.load(f)

def build_xlsx(all_data, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dataset"

    # Header
    headers = ["id", "query", "translated_query", "intent", "language"]
    header_fill = PatternFill("solid", fgColor="2E4057")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 12

    # Dropdown validation on intent column (D2:D301)
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(INTENTS)}"',
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid intent",
        error=f"Choose one of: {', '.join(INTENTS)}",
    )
    ws.add_data_validation(dv)
    dv.sqref = f"D2:D{len(all_data)+1}"

    # Rows
    for row_idx, entry in enumerate(all_data, 2):
        lang = entry["language"]
        fill = PatternFill("solid", fgColor=LANG_COLORS.get(lang, "FFFFFF"))
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col, value=entry[key])
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=(col in (2, 3)))

    # Freeze header
    ws.freeze_panes = "A2"

    # Legend sheet
    lg = wb.create_sheet("Legend")
    lg.column_dimensions["A"].width = 20
    lg.column_dimensions["B"].width = 40
    lg.append(["Language", "Color"])
    for lang, color in LANG_COLORS.items():
        cell_a = lg.cell(row=lg.max_row+1, column=1, value=lang)
        cell_b = lg.cell(row=lg.max_row, column=2, value=f"#{color}")
        cell_a.fill = PatternFill("solid", fgColor=color)
        cell_b.fill = PatternFill("solid", fgColor=color)
    lg.append([])
    lg.append(["Valid intents", ""])
    for intent in INTENTS:
        lg.append([intent, ""])

    wb.save(path)
    print(f"Saved {path} with {len(all_data)} rows")

if __name__ == "__main__":
    existing = load_existing()
    all_data = existing + NEW_ENTRIES
    assert len(all_data) == 300, f"Expected 300, got {len(all_data)}"

    # Save updated JSON
    with open("intent_classification_dataset.json", "w", encoding="utf-8") as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)

    # Save XLSX
    build_xlsx(all_data, "intent_classification_dataset.xlsx")

    # Quick distribution check
    from collections import Counter
    intents = Counter(d["intent"] for d in all_data)
    langs   = Counter(d["language"] for d in all_data)
    print("\nIntent distribution:", dict(intents))
    print("Language distribution:", dict(langs))
